"""엑셀 파일을 TSV 텍스트로 변환하는 모듈."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _cell_to_str(value: Any) -> str:
    """셀 값을 문자열로 변환. None은 빈 문자열."""
    if value is None:
        return ""
    if isinstance(value, float):
        # 소수점 불필요한 .0 제거
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def sheet_to_tsv(ws: Worksheet, max_rows: int | None = None) -> str:
    """워크시트를 TSV 문자열로 변환.

    Args:
        ws: openpyxl Worksheet
        max_rows: 최대 행 수 (None이면 전체)

    Returns:
        TSV 형식 문자열. 첫 열은 행 번호.
    """
    lines: list[str] = []
    row_count = 0
    row_idx = 0  # read_only 모드에서 EmptyCell 대응용
    for row in ws.iter_rows(min_row=1, values_only=False):
        row_idx += 1
        values = [_cell_to_str(cell.value) for cell in row]
        # 완전히 빈 행은 건너뛰기
        if all(v == "" for v in values):
            continue
        # read_only 모드에서 EmptyCell은 .row 속성이 없을 수 있음
        try:
            row_num = row[0].row
        except AttributeError:
            row_num = row_idx
        line = f"{row_num}\t" + "\t".join(values)
        lines.append(line)
        row_count += 1
        if max_rows and row_count >= max_rows:
            break
    return "\n".join(lines)


def parse_excel(file_bytes: bytes) -> dict[str, dict]:
    """엑셀 파일을 파싱하여 시트별 정보를 반환.

    Args:
        file_bytes: 엑셀 파일의 바이트 데이터

    Returns:
        {
            "시트이름": {
                "preview": TSV 문자열 (처음 20행),
                "full": TSV 문자열 (전체),
                "row_count": 전체 행 수,
                "col_count": 전체 열 수,
            }
        }
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheets: dict[str, dict] = {}

    for name in wb.sheetnames:
        ws = wb[name]
        # 전체 TSV
        full_tsv = sheet_to_tsv(ws)
        # 미리보기 (10행) — 토큰 절약
        preview_tsv = sheet_to_tsv(ws, max_rows=10)

        # 행/열 수 계산
        full_lines = [l for l in full_tsv.split("\n") if l.strip()]
        row_count = len(full_lines)
        col_count = max((len(l.split("\t")) for l in full_lines), default=0) if full_lines else 0

        sheets[name] = {
            "preview": preview_tsv,
            "full": full_tsv,
            "row_count": row_count,
            "col_count": col_count,
        }

    wb.close()
    return sheets


def get_structure_summary(sheets: dict[str, dict], file_name: str = "") -> str:
    """시트 구조 요약 텍스트 생성."""
    lines = ["=== 엑셀 구조 요약 ==="]
    if file_name:
        lines.append(f"파일명: {file_name}")
    for name, info in sheets.items():
        lines.append(f"\n[시트: {name}] ({info['row_count']}행 x {info['col_count']}열)")
        lines.append(info["preview"])
    return "\n".join(lines)


def get_full_sheet_data(sheets: dict[str, dict], sheet_names: list[str]) -> str:
    """지정된 시트의 전체 데이터를 반환. trimmed 버전이 있으면 우선 사용."""
    lines = []
    for name in sheet_names:
        if name in sheets:
            # trimmed 버전이 있으면 (토큰 한도 초과 시) 사용
            data = sheets[name].get("full_trimmed", sheets[name]["full"])
            lines.append(f"=== 시트: {name} (전체 데이터) ===")
            lines.append(data)
            lines.append("")
    return "\n".join(lines)
